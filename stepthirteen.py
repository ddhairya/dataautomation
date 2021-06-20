import openpyxl as xl
from openpyxl.worksheet.datavalidation import DataValidation
import master
from pathlib import Path
import os

step = "Thirteen"
status = ""

if Path(master.finalpath + "/" + master.cur_month + master.cur_year + "/tm_exception.xlsx").exists():
    print("Step " + step)
    if not Path(master.finalpath + master.cur_month + master.cur_year + "/TM").exists():
        os.mkdir(master.finalpath + master.cur_month + master.cur_year + "/TM/")

    dv = DataValidation(type="list", formula1='"WH, OT, M, H, LWP, Al, SL"', allow_blank=True)
    wb1 = xl.load_workbook(master.finalpath + "/" + master.cur_month + master.cur_year + "/tm_exception.xlsx")
    sh1 = wb1["Sheet"]
    exp_maxrow = sh1.max_row

    wb2 = xl.load_workbook(master.masterpath + "tmmaster.xlsx")
    sh2 = wb2["TM"]
    tm_maxrow = sh2.max_row

    for tm_row in range(1, tm_maxrow+1):

        wb3 = xl.Workbook()
        sh3 = wb3["Sheet"]
        tmrow = 1

        remark1_cell = sh3.cell(1, 10)
        remark2_cell = sh3.cell(2, 10)
        remark3_cell = sh3.cell(3, 10)
        remark4_cell = sh3.cell(4, 10)
        remark5_cell = sh3.cell(5, 10)
        remark6_cell = sh3.cell(6, 10)
        remark7_cell = sh3.cell(7, 10)

        remark1_cell.value = "WH	=	Weekoff"
        remark2_cell.value = "OT  =  OverTime"
        remark3_cell.value = "M = Miss Punch"
        remark4_cell.value = "H	=	Public Holiday"
        remark5_cell.value = "LWP	=	Leave Without Pay"
        remark6_cell.value = "AL	=	Leave with Pay"
        remark7_cell.value = "SL	=	Sick Leave"



        tm_name = sh2.cell(tm_row, 1).value

        # print(tm_name)

        for row in range(1, exp_maxrow+1):

            empid_cell_val = sh1.cell(row, 1).value
            empname_cell_val = sh1.cell(row, 2).value
            tmname_cell_val = sh1.cell(row, 3).value
            saldate_cell_val = sh1.cell(row, 4).value
            loca_cell_val = sh1.cell(row, 5).value
            tot_hrs_cell_val = sh1.cell(row, 6).value
            remark_cell_val = sh1.cell(row, 7).value

            empid_cell = sh3.cell(tmrow, 1)
            empname_cell = sh3.cell(tmrow, 2)
            tmname_cell = sh3.cell(tmrow, 3)
            saldate_cell = sh3.cell(tmrow, 4)
            loca_cell = sh3.cell(tmrow, 5)
            tot_hrs_cell = sh3.cell(tmrow, 6)
            remark_cell = sh3.cell(tmrow, 7)

            tm = sh1.cell(row, 3).value

            if tmname_cell_val == tm_name or row == 1:
                if row != 1:
                    dv.add(f'G{tmrow}')
                empid_cell.value = empid_cell_val
                empname_cell.value = empname_cell_val
                tmname_cell.value = tmname_cell_val
                saldate_cell.value = saldate_cell_val
                loca_cell.value = loca_cell_val
                tot_hrs_cell.value = tot_hrs_cell_val
                remark_cell.value = remark_cell_val

                tmrow += 1
        sh3.add_data_validation(dv)
        wb3.save(master.finalpath  + master.cur_month + master.cur_year + '/' + "TM/" + tm_name+'.xlsx')

    status = "Done"

