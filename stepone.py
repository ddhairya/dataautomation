from pathlib import Path
import openpyxl as xl
import datelist
import os

import master

step = "one"
status = ""
# making the dir if path not exist
path = Path("c:/it/PayrollAuto/")
if path.exists():
    # making temp directory if not exist
    if not Path(master.temppath).exists():
        os.mkdir(master.temppath)
    # checking the file exists to do automation
    if master.file.exists():
        # load the master file
        print("Step " + step)
        wb = xl.load_workbook(master.filepath+"empshiftsum.xlsx")
        sh1 = wb["Sheet1"]
        sh2 = wb.create_sheet("stepone")
        maxrow = sh1.max_row
        # print(maxrow)
        # loop through each row
        for row in range(1,maxrow+1):
            raw_data = sh1.cell(row,1).value
            raw_wrhr = sh1.cell(row,7).value
            # it will store the working hours in new sheet
            # print(row)
            # print(type(raw_wrhr))
            if isinstance(raw_wrhr,float) or isinstance(raw_wrhr,int):
                new_raw_wrhr = sh2.cell(row,5)
                new_raw_wrhr.value = raw_wrhr
            # it will check the rows with date and overcome the exception is it's run in the month of Jan
            # if it's not date than it means it's either company or employee
            if not raw_data.find(str(datelist.curt_year)) != -1 and not raw_data.find(str(datelist.curt_year-1)) != -1:
                # company
                if raw_data.upper().find(master.company) != -1:
                    new_raw_comp = sh2.cell(row,4)
                    new_raw_comp.value = raw_data
                    new_raw_compcode = sh2.cell(row,1)
                    new_raw_compcode.value = master.company_code
                # employee
                else:
                    # print(raw_data)
                    new_raw_emp = sh2.cell(row,2)
                    new_raw_emp.value = raw_data
                    new_raw_compcode = sh2.cell(row,1)
                    new_raw_compcode.value = master.company_code
            else:

                new_raw_compcode = sh2.cell(row, 1)
                new_raw_compcode.value = master.company_code
                # removing empty rows and taking only date
                if isinstance(raw_wrhr,float) or isinstance(raw_wrhr,int):
                    new_raw_date = sh2.cell(row,3)
                    new_raw_date.value = raw_data
        # saving in the workbook
        wb.save(master.temppath + "stepone.xlsx")
        status = "Done"
    else:
        print("Please copy the empshiftsum.xlsx inside c:/it/PayrollAuto/")
else:
    path = "c:/it/PayrollAuto/"
    os.mkdir(path)


