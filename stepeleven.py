import os
import openpyxl as xl
import master
import datelist
from pathlib import Path
import datetime
from shutil import copyfile, rmtree

step = "Eleven"
status = ""


def nonValue(field):
    if field is None:
        return 0
    else:
        return field


if Path(master.temppath+"payroll.xlsx").exists():
    print("Step "+ step)
    if datelist.cur_numdays > datelist.numdays:
        wb1 = xl.load_workbook(master.tempfile + "payroll.xlsx")
        sh1 = wb1["Sheet"]
        payroll_maxrow = sh1.max_row

        wb3 = xl.Workbook()
        sh3 = wb3["Sheet"]

        for row in range(1, payroll_maxrow + 1):    
            cmpycode_val = sh1.cell(row, 1).value
            empcode_val = sh1.cell(row, 2).value
            SalaryDates_val = sh1.cell(row, 3).value
            SalaryDay = SalaryDates_val.split("/")
            Nhrs_val = sh1.cell(row, 4).value
            N_OT_HRS_val = sh1.cell(row, 5).value
            H_Othrs_val = sh1.cell(row, 6).value
            EXTRA_EXT_OT_val = sh1.cell(row, 7).value
            W_OT_hrs_val = sh1.cell(row, 8).value
            Weekoff_val = sh1.cell(row, 9).value
            SiteID_val = sh1.cell(row, 10).value
            Costcode_val = sh1.cell(row, 11).value

            cmpycode_cell = sh3.cell(row, 1)
            empcode_cell = sh3.cell(row, 2)
            SalaryDates_cell = sh3.cell(row, 3)
            Nhrs_cell = sh3.cell(row, 4)
            N_OT_HRS_cell = sh3.cell(row, 5)
            H_Othrs_cell = sh3.cell(row, 6)
            EXTRA_EXT_OT_cell = sh3.cell(row, 7)
            W_OT_hrs_cell = sh3.cell(row, 8)
            Weekoff_cell = sh3.cell(row, 9)
            SiteID_cell = sh3.cell(row, 10)
            Costcode_cell = sh3.cell(row, 11)
            
            if row > 1 and int(SalaryDay[0]) <= datelist.cur_numdays:
                NewSalaryDate_val = SalaryDay[0] + "/" + '{:02d}'.format(datelist.curt_month) + "/" + str(datelist.curt_year)
                SalaryDates_val = NewSalaryDate_val
            
            cmpycode_cell.value = cmpycode_val
            empcode_cell.value = empcode_val
            SalaryDates_cell.value = SalaryDates_val
            Nhrs_cell.value = Nhrs_val
            N_OT_HRS_cell.value = N_OT_HRS_val
            H_Othrs_cell.value = H_Othrs_val
            EXTRA_EXT_OT_cell.value = EXTRA_EXT_OT_val
            W_OT_hrs_cell.value = W_OT_hrs_val
            Weekoff_cell.value = Weekoff_val
            SiteID_cell.value = SiteID_val
            Costcode_cell.value = Costcode_val

            filtercell = "A1:T" + str(payroll_maxrow + 1)
            sh3.auto_filter.ref = filtercell

            wb3.save(master.finalpath + "/" + master.cur_month + master.cur_year + "/payroll.xlsx")
    else:
        wb1 = xl.load_workbook(master.tempfile + "payroll.xlsx")
        sh1 = wb1["Sheet"]
        payroll_maxrow = sh1.max_row

        wb3 = xl.Workbook()
        sh3 = wb3["Sheet"]
        payroll_row = 1

        wb4 = xl.Workbook()
        sh4 = wb4["Sheet"]
        month_exception_row = 1

        # last_day = datetime.datetime.today().replace(day=datelist.cur_numdays)

        for row in range(1,payroll_maxrow+1):
            cmpycode_val = sh1.cell(row, 1).value
            empcode_val = sh1.cell(row, 2).value
            SalaryDates_val = sh1.cell(row, 3).value
            SalaryDay = SalaryDates_val.split("/")
            Nhrs_val = sh1.cell(row, 4).value
            N_OT_HRS_val = sh1.cell(row, 5).value
            H_Othrs_val = sh1.cell(row, 6).value
            EXTRA_EXT_OT_val = sh1.cell(row, 7).value
            W_OT_hrs_val = sh1.cell(row, 8).value
            Weekoff_val = sh1.cell(row, 9).value
            SiteID_val = sh1.cell(row, 10).value
            Costcode_val = sh1.cell(row, 11).value

            # print(type(SalaryDates_val))
            # print(type(last_day))

            cmpycode_cell = sh3.cell(payroll_row, 1)
            empcode_cell = sh3.cell(payroll_row, 2)
            SalaryDates_cell = sh3.cell(payroll_row, 3)
            Nhrs_cell = sh3.cell(payroll_row, 4)
            N_OT_HRS_cell = sh3.cell(payroll_row, 5)
            H_Othrs_cell = sh3.cell(payroll_row, 6)
            EXTRA_EXT_OT_cell = sh3.cell(payroll_row, 7)
            W_OT_hrs_cell = sh3.cell(payroll_row, 8)
            Weekoff_cell = sh3.cell(payroll_row, 9)
            SiteID_cell = sh3.cell(payroll_row, 10)
            Costcode_cell = sh3.cell(payroll_row, 11)

            cmpycode_cell_month_exp = sh4.cell(month_exception_row, 1)
            empcode_cell_month_exp = sh4.cell(month_exception_row, 2)
            SalaryDates_cell_month_exp = sh4.cell(month_exception_row, 3)
            Nhrs_cell_month_exp = sh4.cell(month_exception_row, 4)
            N_OT_HRS_cell_month_exp = sh4.cell(month_exception_row, 5)
            H_Othrs_cell_month_exp = sh4.cell(month_exception_row, 6)
            EXTRA_EXT_OT_cell_month_exp = sh4.cell(month_exception_row, 7)
            W_OT_hrs_cell_month_exp = sh4.cell(month_exception_row, 8)
            Weekoff_cell_month_exp = sh4.cell(month_exception_row, 9)
            SiteID_cell_month_exp = sh4.cell(month_exception_row, 10)
            Costcode_cell_month_exp = sh4.cell(month_exception_row, 11)

            if row == 1:
                cmpycode_cell.value = cmpycode_val
                empcode_cell.value = empcode_val
                SalaryDates_cell.value = SalaryDates_val
                Nhrs_cell.value = Nhrs_val
                N_OT_HRS_cell.value = N_OT_HRS_val
                H_Othrs_cell.value = H_Othrs_val
                EXTRA_EXT_OT_cell.value = EXTRA_EXT_OT_val
                W_OT_hrs_cell.value = W_OT_hrs_val
                Weekoff_cell.value = Weekoff_val
                SiteID_cell.value = SiteID_val
                Costcode_cell.value = Costcode_val

                cmpycode_cell_month_exp.value = cmpycode_val
                empcode_cell_month_exp.value = empcode_val
                SalaryDates_cell_month_exp.value = SalaryDates_val
                Nhrs_cell_month_exp.value = Nhrs_val
                N_OT_HRS_cell_month_exp.value = N_OT_HRS_val
                H_Othrs_cell_month_exp.value = H_Othrs_val
                EXTRA_EXT_OT_cell_month_exp.value = EXTRA_EXT_OT_val
                W_OT_hrs_cell_month_exp.value = W_OT_hrs_val
                Weekoff_cell_month_exp.value = Weekoff_val
                SiteID_cell_month_exp.value = SiteID_val
                Costcode_cell_month_exp.value = Costcode_val

                month_exception_row += 1
                payroll_row += 1
                
            elif row > 1 and int(SalaryDay[0]) <= datelist.cur_numdays:
                NewSalaryDate_val = SalaryDay[0] + "/" + '{:02d}'.format(datelist.curt_month) + "/" + str(datelist.curt_year)
                SalaryDates_val = NewSalaryDate_val

                cmpycode_cell.value = cmpycode_val
                empcode_cell.value = empcode_val
                SalaryDates_cell.value = SalaryDates_val
                Nhrs_cell.value = Nhrs_val
                N_OT_HRS_cell.value = nonValue(N_OT_HRS_val)
                H_Othrs_cell.value = nonValue(H_Othrs_val)
                EXTRA_EXT_OT_cell.value = nonValue(EXTRA_EXT_OT_val)
                W_OT_hrs_cell.value = nonValue(W_OT_hrs_val)
                Weekoff_cell.value = Weekoff_val
                SiteID_cell.value = SiteID_val
                Costcode_cell.value = Costcode_val
                payroll_row += 1
            elif row > 1 and not int(SalaryDay[0]) <= datelist.cur_numdays:
                if Nhrs_val == 9 and N_OT_HRS_val == 0:
                    pass
                else:
                    cmpycode_cell_month_exp.value = cmpycode_val
                    empcode_cell_month_exp.value = empcode_val
                    SalaryDates_cell_month_exp.value = SalaryDates_val
                    Nhrs_cell_month_exp.value = Nhrs_val
                    N_OT_HRS_cell_month_exp.value = N_OT_HRS_val
                    H_Othrs_cell_month_exp.value = H_Othrs_val
                    EXTRA_EXT_OT_cell_month_exp.value = EXTRA_EXT_OT_val
                    W_OT_hrs_cell_month_exp.value = W_OT_hrs_val
                    Weekoff_cell_month_exp.value = Weekoff_val
                    SiteID_cell_month_exp.value = SiteID_val
                    Costcode_cell_month_exp.value = Costcode_val
                    month_exception_row += 1

        # wb3.save(master.temppath + "/month_payroll.xlsx")

        filtercell = "A1:T" + str(payroll_maxrow + 1)
        sh3.auto_filter.ref = filtercell

        wb3.save(master.finalpath + "/" + master.cur_month + master.cur_year + "/payroll.xlsx")

        wb4.save(master.finalpath + "/" + master.cur_month + master.cur_year  + "/month_exception.xlsx")
        rmtree(master.temppath)
        status = "Done"

