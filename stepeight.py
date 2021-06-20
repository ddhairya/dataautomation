import master
import os
import openpyxl as xl
from pathlib import Path

# taking the master and based on the department and punch assigning the cost center
step = "eight"
status = ""

if Path(master.masterpath+"costcenter.xlsx").exists() and Path(master.temppath+"payroll_department.xlsx").exists():
    print("Step "+step)
    wb1 = xl.load_workbook(master.masterfile+"costcenter.xlsx")
    sh1 = wb1["Sheet1"]
    maxmasterrow = sh1.max_row
    wb2 = xl.load_workbook(master.tempfile+"payroll_department.xlsx")
    sh2 = wb2["Sheet"]
    maxrow = sh2.max_row

    # create new workbook with default Sheet
    wb3 = xl.Workbook()
    sh3 = wb3["Sheet"]
    deleted_data_row = 2
    
    # print(maxrow)

    cort_compcode_data_cell = sh2.cell(1, 1)
    cort_emp_data_cell = sh2.cell(1, 2)
    cort_date_data_cell = sh2.cell(1, 3)
    cort_comp_data_cell = sh2.cell(1, 20)
    cort_wrkhrs_data_cell = sh2.cell(1, 4)
    cort_not_data_cell = sh2.cell(1, 5)
    cort_hot_data_cell = sh2.cell(1, 6)
    cort_exot_data_cell = sh2.cell(1, 7)
    cort_wot_data_cell = sh2.cell(1, 8)
    cort_weekoff_data_cell = sh2.cell(1, 9)
    cort_siteid_data_cell = sh2.cell(1, 10)
    cort_costcode_data_cell = sh2.cell(1, 11)

    cort_compcode_data_cell.value = "cmpycode"
    cort_emp_data_cell.value = "empcode"
    cort_date_data_cell.value = "SalaryDates"
    cort_comp_data_cell.value = "Comp"
    cort_wrkhrs_data_cell.value = "Nhrs"
    cort_not_data_cell.value = "N_ OT_HRS"
    cort_hot_data_cell.value = "H_Othrs"
    cort_exot_data_cell.value = "EXTRA EXT OT"
    cort_wot_data_cell.value = "W_OT_hrs"
    cort_weekoff_data_cell.value = "Weekoff"
    cort_siteid_data_cell.value = "SiteID"
    cort_costcode_data_cell.value = "Costcode"

    del_emp_data_cell = sh3.cell(1, 1)
    del_date_data_cell = sh3.cell(1, 2)
    del_comp_data_cell = sh3.cell(1, 3)

    del_emp_data_cell.value = "empcode"
    del_date_data_cell.value = "SalaryDates"
    del_comp_data_cell.value = "Comp"
    
    for row in range(2,maxrow+1):
        cort_compcode_data_cell = sh2.cell(row, 1)
        cort_date_data_cell = sh2.cell(row, 3)
        cort_wrkhrs_data_cell = sh2.cell(row, 4)
        cort_not_data_cell = sh2.cell(row, 5)
        cort_hot_data_cell = sh2.cell(row, 6)
        cort_exot_data_cell = sh2.cell(row, 7)
        cort_wot_data_cell = sh2.cell(row, 8)
        cort_weekoff_data_cell = sh2.cell(row, 9)
        cort_siteid_data_cell = sh2.cell(row, 10)
        cort_costcode_data_cell = sh2.cell(row, 11)
        cort_depart_data_cell = sh2.cell(row, 19)
        cort_comp_data_cell = sh2.cell(row, 20)

        emp_data_cell = sh2.cell(row, 2)
        siteid_data_cell = sh2.cell(row, 10)
        costcenter_data_cell = sh2.cell(row,11)

        emp_data = sh2.cell(row,2).value
        department_data = sh2.cell(row,19).value
        comp_data = sh2.cell(row,20).value
        date_data = sh2.cell(row,3).value
        # print(department_data )
        # print(comp_data)


        del_emp_data_cell = sh3.cell(deleted_data_row, 1)
        del_date_data_cell = sh3.cell(deleted_data_row, 2)
        del_comp_data_cell = sh3.cell(deleted_data_row, 3)

        for masterrow in range(2,maxmasterrow+1):
            mastercomp_data = str(sh1.cell(masterrow,1).value)
            bakery_data = str(sh1.cell(masterrow,2).value)
            pastry_data = str(sh1.cell(masterrow, 3).value)
            bar_data = str(sh1.cell(masterrow, 4).value)
            mainstore_data = str(sh1.cell(masterrow, 5).value)
            kitchen_data = str(sh1.cell(masterrow, 6).value)

            if emp_data is not None and emp_data[:-4] != master.emp_start_code or department_data is None or comp_data == "LA BRIOCHE CPU":
                # print(row)
                # print(emp_data + date_data)
                # print(emp_data[:-4])

                emp_data_cell.value = ""
                cort_compcode_data_cell.value = ""
                cort_date_data_cell.value = ""
                cort_comp_data_cell.value = ""
                cort_wrkhrs_data_cell.value = ""
                cort_not_data_cell.value = ""
                cort_hot_data_cell.value = ""
                cort_exot_data_cell.value = ""
                cort_wot_data_cell.value = ""
                cort_weekoff_data_cell.value = ""
                cort_siteid_data_cell.value = ""
                cort_costcode_data_cell.value = ""
                cort_depart_data_cell.value = ""

                # print(row)
                # print("Delete")
                # row = row - 1
                # delete the row where emp code is with Z
                # sh2.delete_rows(row, 1)

                del_emp_data_cell.value = emp_data
                del_date_data_cell.value = date_data
                del_comp_data_cell.value = comp_data

                deleted_data_row += 1
                break

            elif comp_data == mastercomp_data and department_data is not None and bakery_data.find(department_data) != -1:
                siteid_data_cell.value = bakery_data[-3:]
                costcenter_data_cell.value = bakery_data[-3:]
                # print( emp_data + "-----" + comp_data + "----------" + department_data + "----------------" + siteid_data_cell.value)
                break

            elif comp_data == mastercomp_data and department_data is not None and pastry_data.find(department_data) != -1:
                siteid_data_cell.value = pastry_data[-3:]
                costcenter_data_cell.value = pastry_data[-3:]
                # print( emp_data + "-----" + comp_data + "----------" + department_data + "----------------" + siteid_data_cell.value)
                break

            elif comp_data == mastercomp_data and department_data is not None and bar_data.find(department_data) != -1:
                siteid_data_cell.value = bar_data[-3:]
                costcenter_data_cell.value = bar_data[-3:]
                # print( emp_data + "-----" + comp_data + "----------" + department_data + "----------------" + siteid_data_cell.value)
                break

            elif comp_data == mastercomp_data and department_data is not None and mainstore_data.find(department_data) != -1:
                siteid_data_cell.value = mainstore_data[-3:]
                costcenter_data_cell.value = mainstore_data[-3:]
                # print( emp_data + "-----" + comp_data + "----------" + department_data + "----------------" + siteid_data_cell.value)
                break

            elif comp_data == mastercomp_data and department_data is not None and kitchen_data.find(department_data) != -1:
                siteid_data_cell.value = kitchen_data[-3:]
                costcenter_data_cell.value = kitchen_data[-3:]
                # print( emp_data + "-----" + comp_data + "----------" + department_data + "----------------" + siteid_data_cell.value)
                break

            elif comp_data == mastercomp_data and department_data is not None and department_data == "ADMIN":
                siteid_data_cell.value = "481"
                costcenter_data_cell.value = "481"
                # print( emp_data + "-----" + comp_data + "----------" + department_data + "----------------" + siteid_data_cell.value)
                break

    wb2.save(master.temppath+"payroll.xlsx")
    wb3.save(master.temppath+"delete.xlsx")
    status = "Done"
