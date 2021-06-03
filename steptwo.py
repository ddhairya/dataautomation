from pathlib import  Path
import openpyxl as xl

import master
import stepone

step = "two"
status = ""
if Path(master.temppath + "stepone.xlsx").exists():
    print("Step "+step)
    wb = xl.load_workbook(master.tempfile+"stepone.xlsx")
    sh1 = wb["stepone"]
    sh2 = wb.create_sheet("steptwo")
    temp_comp_data = ""
    temp_emp_data = ""
    maxrow = sh1.max_row

    for row in range(1,maxrow+1):
        compcode_data = sh1.cell(row, 1).value
        emp_data = sh1.cell(row, 2).value
        date_data = sh1.cell(row,3).value
        comp_data = sh1.cell(row, 4).value
        wrkhrs_data = sh1.cell(row,5).value

        new_compcode_data_cell = sh2.cell(row, 1)
        new_emp_data_cell = sh2.cell(row,2)
        new_date_data_cell = sh2.cell(row,3)
        new_comp_data_cell = sh2.cell(row, 4)
        new_wrkhrs_data_cell = sh2.cell(row,5)
        # if condition try to get the compy name next to each records
        if comp_data is not None:
            temp_comp_data = comp_data
        # elif date_data is not None:
        else:
            new_comp_data_cell.value = temp_comp_data
        # if condition try to get the emp name next to each records
        if emp_data is not None:
            temp_emp_data = emp_data
            # print(temp_emp_data)
            # this will help to get into next step
            new_emp_data_cell.value = temp_emp_data
            new_compcode_data_cell.value = compcode_data
        else:
            new_emp_data_cell.value = temp_emp_data
            new_compcode_data_cell.value = compcode_data
        new_date_data_cell.value = date_data
        new_wrkhrs_data_cell.value = wrkhrs_data
    wb.save(master.temppath+"steptwo.xlsx")
    status = "Done"



