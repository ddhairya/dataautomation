import master
import openpyxl as xl
from pathlib import Path
import os
from shutil import copyfile, rmtree

step = "ten"
status = ""


if Path(master.temppath+"payroll.xlsx").exists() and Path(master.temppath+"exception.xlsx").exists():
    print("Step " + step)

    wb1 = xl.load_workbook(master.tempfile+"payroll.xlsx")
    sh1 = wb1["Sheet"]
    payroll_maxrow = sh1.max_row

    wb2 = xl.load_workbook(master.tempfile + "exception.xlsx")
    sh2 = wb2["Sheet"]
    exception_maxroww = sh2.max_row

    wb3 = xl.Workbook()
    sh3 = wb3["Sheet"]
    payroll_row = 1
    
    wb4 = xl.Workbook()
    sh4 = wb4["Sheet"]
    exception_row = 1

    wb5 = xl.load_workbook(master.finalpath+"/"+master.cur_month+master.cur_year+"/ignored.xlsx")
    sh5 = wb5["Sheet"]
    ignore_max = sh5.max_row
    # print(ignore_max)

    for row in range(1,payroll_maxrow):
        cmpycode_val = sh1.cell(row,1).value
        empcode_val = sh1.cell(row,2).value
        SalaryDates_val = sh1.cell(row,3).value
        Nhrs_val = sh1.cell(row,4).value
        N_OT_HRS_val = sh1.cell(row,5).value
        H_Othrs_val = sh1.cell(row,6).value
        EXTRA_EXT_OT_val = sh1.cell(row,7).value
        W_OT_hrs_val = sh1.cell(row,8).value
        Weekoff_val = sh1.cell(row,9).value
        SiteID_val = sh1.cell(row,10).value
        Costcode_val = sh1.cell(row,11).value

        cmpycode_cell = sh3.cell(payroll_row,1)
        empcode_cell = sh3.cell(payroll_row,2)
        SalaryDates_cell = sh3.cell(payroll_row,3)
        Nhrs_cell = sh3.cell(payroll_row,4)
        N_OT_HRS_cell = sh3.cell(payroll_row,5)
        H_Othrs_cell = sh3.cell(payroll_row,6)
        EXTRA_EXT_OT_cell = sh3.cell(payroll_row,7)
        W_OT_hrs_cell = sh3.cell(payroll_row,8)
        Weekoff_cell = sh3.cell(payroll_row,9)
        SiteID_cell = sh3.cell(payroll_row,10)
        Costcode_cell = sh3.cell(payroll_row,11)

        if empcode_val is not None and Costcode_val is None:
            ig_emp_data_cell = sh5.cell(ignore_max, 1)
            ig_date_data_cell = sh5.cell(ignore_max, 2)
            ig_comp_data_cell = sh5.cell(ignore_max, 3)

            ig_date_data_cell.value = SalaryDates_val
            ig_emp_data_cell.value = empcode_val
            # print(ignore_max)
            ignore_max += 1
        elif empcode_val is not None and Costcode_val is not None:
            cmpycode_cell.value = cmpycode_val
            empcode_cell.value = empcode_val
            SalaryDates_cell.value = SalaryDates_val
            Nhrs_cell.value = Nhrs_val
            N_OT_HRS_cell.value = N_OT_HRS_val
            H_Othrs_cell.value = H_Othrs_val
            EXTRA_EXT_OT_cell.value = EXTRA_EXT_OT_val
            W_OT_hrs_cell.value = W_OT_hrs_val
            Weekoff_cell.value = Weekoff_val
            SiteID_cell.value = SiteID_val
            Costcode_cell.value = Costcode_val

            payroll_row += 1



    if not Path(master.finalpath + master.cur_month + master.cur_year).exists():
        os.mkdir(master.finalpath + master.cur_month + master.cur_year)

    # filtercell = "A1:T" + str(payroll_maxrow + 1)
    # sh3.auto_filter.ref = filtercell
    # wb3.save(master.finalpath + "/" + master.cur_month + master.cur_year + "/payroll.xlsx")
    wb3.save(master.temppath + "/payroll.xlsx")
    wb5.save(master.finalpath + "/" + master.cur_month + master.cur_year + "/ignored.xlsx")
    # wb3.save(master.finalpath+"payroll.xlsx")

    for row in range(1,exception_maxroww):
        cmpycode_val = sh2.cell(row,1).value
        empcode_val = sh2.cell(row,2).value
        SalaryDates_val = sh2.cell(row,3).value
        Nhrs_val = sh2.cell(row,4).value
        N_OT_HRS_val = sh2.cell(row,5).value
        H_Othrs_val = sh2.cell(row,6).value
        EXTRA_EXT_OT_val = sh2.cell(row,7).value
        W_OT_hrs_val = sh2.cell(row,8).value
        Weekoff_val = sh2.cell(row,9).value
        SiteID_val = sh2.cell(row,10).value
        Costcode_val = sh2.cell(row,11).value
       
        cmpycode_cell = sh4.cell(exception_row,1)
        empcode_cell = sh4.cell(exception_row,2)
        SalaryDates_cell = sh4.cell(exception_row,3)
        Nhrs_cell = sh4.cell(exception_row,4)
        N_OT_HRS_cell = sh4.cell(exception_row,5)
        H_Othrs_cell = sh4.cell(exception_row,6)
        EXTRA_EXT_OT_cell = sh4.cell(exception_row,7)
        W_OT_hrs_cell = sh4.cell(exception_row,8)
        Weekoff_cell = sh4.cell(exception_row,9)
        SiteID_cell = sh4.cell(exception_row,10)
        Costcode_cell = sh4.cell(exception_row,11)
        
        if empcode_val is not None:
            cmpycode_cell.value = cmpycode_val
            empcode_cell.value = empcode_val
            SalaryDates_cell.value = SalaryDates_val
            Nhrs_cell.value = Nhrs_val
            N_OT_HRS_cell.value = N_OT_HRS_val
            H_Othrs_cell.value = H_Othrs_val
            EXTRA_EXT_OT_cell.value = EXTRA_EXT_OT_val
            W_OT_hrs_cell.value = W_OT_hrs_val
            Weekoff_cell.value = Weekoff_val
            SiteID_cell.value = SiteID_val
            Costcode_cell.value = Costcode_val
        
            exception_row += 1

    if not Path(master.finalpath + master.cur_month + master.cur_year).exists():
        os.mkdir(master.finalpath + master.cur_month + master.cur_year)

    filtercell = "A1:T" + str(exception_maxroww + 1)
    sh4.auto_filter.ref = filtercell

    wb4.save(master.finalpath + "/" + master.cur_month + master.cur_year + "/exception.xlsx")

    # wb4.save(master.temppath + "/exception.xlsx")
    copyfile(master.finalpath + "empshiftsum.xlsx",
             master.finalpath + "/" + master.cur_month + master.cur_year + "/empshiftsum.xlsx")
    # rmtree(master.temppath)
    status = "Done"
    # wb4.save(master.finalpath+"exception.xlsx")