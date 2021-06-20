import os
import openpyxl as xl
import master
from pathlib import Path


step = "Twelve"
status = ""

if Path(master.finalpath + "/" + master.cur_month + master.cur_year + "/exception.xlsx").exists() and Path(master.masterpath+"tmmaster.xlsx").exists():
    print("Step " + step)
    wb1 = xl.load_workbook(master.finalpath + "/" + master.cur_month + master.cur_year + "/exception.xlsx")
    sh1 = wb1["Sheet"]
    except_maxrow = sh1.max_row
    # print(except_maxrow)

    wb2 = xl.load_workbook(master.masterpath+"tmmaster.xlsx")
    sh2 = wb2["Sheet"]
    outlet_maxrow = sh2.max_row
    # print(outlet_maxrow)

    wb3 = xl.Workbook()
    sh3 = wb3["Sheet"]
    new_exception = 2

    wb4 = xl.load_workbook(master.masterpath + "empmaster.xlsx")
    sh4 = wb4["Sheet1"]
    emp_maxrow = sh4.max_row
    # print(emp_maxrow)
    
    empid_cell = sh3.cell(1, 1)
    empname_cell = sh3.cell(1, 2)
    tmname_cell = sh3.cell(1, 3)
    saldate_cell = sh3.cell(1, 4)
    loca_cell = sh3.cell(1, 5)
    tot_hrs_cell = sh3.cell(1, 6)
    remark_cell = sh3.cell(1, 7)

    empid_cell.value = "Emp Code"
    empname_cell.value = "Emp Name"
    tmname_cell.value = "TM / TC / EC"
    saldate_cell.value = "Salary Date"
    loca_cell.value = "Location"
    tot_hrs_cell.value = "Hrs"
    remark_cell.value = "Remark"
    
    for row in range (2, except_maxrow+1):
        emp_code = str(sh1.cell(row, 2).value)
        sal_date = sh1.cell(row, 3).value
        cost_cen = str(sh1.cell(row,10).value).upper()

        # print(str(emp_code) + "-----------" + str(cost_cen))

        empid_cell = sh3.cell(new_exception, 1)
        empname_cell = sh3.cell(new_exception, 2)
        tmname_cell = sh3.cell(new_exception, 3)
        saldate_cell = sh3.cell(new_exception, 4)
        loca_cell = sh3.cell(new_exception, 5)
        tot_hrs_cell = sh3.cell(new_exception, 6)
        remark_cell = sh3.cell(new_exception, 7)


        empid_cell.value = emp_code
        saldate_cell.value = sal_date
        loca_cell.value = cost_cen

        if sh1.cell(row, 4).value is not None and sh1.cell(row, 5).value is not None and sh1.cell(row, 7).value is not None:
            tot_hrs_cell.value = sh1.cell(row, 4).value + sh1.cell(row, 5).value + sh1.cell(row, 7).value

        for out_row in range (1, outlet_maxrow+1):
            tm_name = sh2.cell(out_row,2).value
            out_cost_cen = str(sh2.cell(out_row,1).value).upper()
            # print("---++++" + str(out_cost_cen))

            if cost_cen == out_cost_cen:
                # print(emp_code + cost_cen + tm_name)
                tmname_cell.value = tm_name
                break


        for emp_row in range(1, emp_maxrow+1):
            em_name = sh4.cell(emp_row,2).value
            em_id = str(sh4.cell(emp_row,1).value)
            # print(emp_code + em_id)
            if emp_code == em_id:
                # print(emp_code + em_name)
                empname_cell.value = em_name
                break
        new_exception += 1

    filtercell = "A1:G" + str(new_exception + 1)
    sh3.auto_filter.ref = filtercell
    wb3.save(master.finalpath + "/" + master.cur_month + master.cur_year + "/tm_exception.xlsx")


            # print(tot_hrs)
        # nwr_hrs = sh1.cell(row,4).value
        # not_hrs = sh1.cell(row,5).value
        # eot_hrs = sh1.cell(row,7).value

    status = "Done"