import master
import openpyxl as xl
from pathlib import Path
import os

step = "seven"
status = ""

if Path(master.masterpath+"empmaster.xlsx").exists():
    print("Step "+step)
    wb1 = xl.load_workbook(master.masterfile+"empmaster.xlsx")
    sh1 = wb1["Sheet1"]
    mastermaxrow = sh1.max_row
    # print(mastermaxrow)
    wb2 = xl.load_workbook(master.tempfile+"payroll.xlsx")
    sh2 = wb2["Sheet"]
    maxrow = sh2.max_row
    for row in range(2,maxrow+2):
        emp_data = sh2.cell(row,2).value
        department_cell = sh2.cell(row,19)
        for masterrow in range(1,mastermaxrow+1):
            master_emp_data = sh1.cell(masterrow,1).value
            master_department = sh1.cell(masterrow,5).value
            if master_department is not None and emp_data == master_emp_data :
                department_cell.value = master_department
                # print(emp_data + "-" + master_department)
                break
    wb3 = xl.load_workbook(master.tempfile + "exception.xlsx")
    sh3 = wb3["Sheet"]
    maxrow = sh3.max_row
    for row in range(2, maxrow + 2):
        emp_data = sh3.cell(row, 2).value
        department_cell = sh3.cell(row, 19)
        for masterrow in range(1, mastermaxrow + 1):
            master_emp_data = sh1.cell(masterrow, 1).value
            master_department = sh1.cell(masterrow, 5).value
            if master_department is not None and emp_data == master_emp_data:
                department_cell.value = master_department
                # print(emp_data + "-" + master_department)
                break

    wb2.save(master.temppath+"payroll_department.xlsx")
    wb3.save(master.temppath + "exception_department.xlsx")
    status = "Done"
