import openpyxl as xl
from pathlib import Path
import master

step = "six"
status = ""

if Path(master.temppath+"stepfive.xlsx").exists() and Path(master.temppath+"payroll.xlsx").exists():
    print("Step "+step)
    wb1 = xl.load_workbook(master.tempfile+"stepfive.xlsx")
    sh1 = wb1["stepfour"]
    maxrowsh1 = sh1.max_row

    wb2 = xl.load_workbook(master.tempfile+"payroll.xlsx")
    sh2 = wb2["Sheet"]
    maxrowsh2 = sh2.max_row

    # exception file
    wb3 = xl.Workbook()
    sh3 = wb3["Sheet"]

    wb4 = xl.Workbook()
    sh4 = wb4["Sheet"]

    miss_not_data_cell = sh3.cell(1, 5)
    miss_hot_data_cell = sh3.cell(1, 6)
    miss_exot_data_cell = sh3.cell(1, 7)
    miss_wot_data_cell = sh3.cell(1, 8)
    miss_weekoff_data_cell = sh3.cell(1, 9)
    miss_siteid_data_cell = sh3.cell(1, 10)
    miss_costcode_data_cell = sh3.cell(1, 11)

    miss_not_data_cell.value = "N_ OT_HRS "
    miss_hot_data_cell.value = "H_Othrs"
    miss_exot_data_cell.value = "EXTRA EXT OT"
    miss_wot_data_cell.value = "W_OT_hrs"
    miss_weekoff_data_cell.value = "Weekoff"
    miss_siteid_data_cell.value = "SiteID"
    miss_costcode_data_cell.value = "Costcode"

    cort_compcode_data_cell = sh4.cell(1, 1)
    cort_emp_data_cell = sh4.cell(1, 2)
    cort_date_data_cell = sh4.cell(1, 3)
    cort_comp_data_cell = sh4.cell(1, 20)
    cort_wrkhrs_data_cell = sh4.cell(1, 4)
    cort_not_data_cell = sh4.cell(1, 5)
    cort_hot_data_cell = sh4.cell(1, 6)
    cort_exot_data_cell = sh4.cell(1, 7)
    cort_wot_data_cell = sh4.cell(1, 8)
    cort_weekoff_data_cell = sh4.cell(1, 9)
    cort_siteid_data_cell = sh4.cell(1, 10)
    cort_costcode_data_cell = sh4.cell(1, 11)

    cort_compcode_data_cell.value = "cmpycode"
    cort_emp_data_cell.value = "empcode"
    cort_date_data_cell.value = "SalaryDates"
    cort_comp_data_cell.value = "Comp"
    cort_wrkhrs_data_cell.value = "Nhrs"
    cort_not_data_cell.value = "N_ OT_HRS"
    cort_hot_data_cell.value = "H_Othrs"
    cort_exot_data_cell.value = "EXTRA EXT OT"
    cort_wot_data_cell.value = "W_OT_hrs"
    cort_weekoff_data_cell.value = "Weekoff"
    cort_siteid_data_cell.value = "SiteID"
    cort_costcode_data_cell.value = "Costcode"

    for row in range(1, maxrowsh1+1):
        compcode_data = sh1.cell(row, 1).value
        emp_data = sh1.cell(row, 2).value
        date_data = sh1.cell(row, 3).value
        comp_data = sh1.cell(row, 4).value
        wrkhrs_data = sh1.cell(row, 5).value

        miss_compcode_data_cell = sh3.cell(row, 1)
        miss_emp_data_cell = sh3.cell(row, 2)
        miss_date_data_cell = sh3.cell(row, 3)
        miss_comp_data_cell = sh3.cell(row, 20)
        miss_wrkhrs_data_cell = sh3.cell(row, 4)
        miss_not_data_cell = sh3.cell(row, 5)
        miss_exot_data_cell = sh3.cell(row, 7)

        miss_compcode_data_cell.value = compcode_data
        miss_emp_data_cell.value = emp_data
        miss_date_data_cell.value = date_data
        miss_comp_data_cell.value = comp_data

        if isinstance(wrkhrs_data,float):

            if int(wrkhrs_data) > 9:
                miss_wrkhrs_data_cell.value = 9
                # miss_not_data_cell.value = int(wrkhrs_data) - 9
                if int(wrkhrs_data) - 9 > 2:
                    miss_not_data_cell.value = 2
                    miss_exot_data_cell.value = int(wrkhrs_data) - 9 - 2
                else:
                    miss_not_data_cell.value = int(wrkhrs_data) - 9

            # there will never be a scenario where wrk hr will be less than 9, still for safer side
            else:
                miss_wrkhrs_data_cell.value = wrkhrs_data
        else:
            miss_wrkhrs_data_cell.value = wrkhrs_data

    cort_row = 1
    for row in range(1, maxrowsh2+1):
        compcode_data = sh2.cell(row,1).value
        emp_data = sh2.cell(row, 2).value
        date_data = sh2.cell(row, 3).value
        comp_data = sh2.cell(row, 4).value
        wrkhrs_data = sh2.cell(row, 5).value
        not_data = sh2.cell(row,6).value

        if date_data is not None:
            cort_compcode_data_cell = sh4.cell(cort_row, 1)
            cort_emp_data_cell = sh4.cell(cort_row, 2)
            cort_date_data_cell = sh4.cell(cort_row, 3)
            cort_comp_data_cell = sh4.cell(cort_row, 20)
            cort_wrkhrs_data_cell = sh4.cell(cort_row, 4)
            cort_not_data_cell = sh4.cell(cort_row,5)
            cort_weekoff_data_cell = sh4.cell(cort_row, 9)
            
            if row > 1 and float(wrkhrs_data) >= 9.0:
                cort_compcode_data_cell.value = compcode_data
                cort_emp_data_cell.value = emp_data
                cort_date_data_cell.value = date_data
                cort_comp_data_cell.value = comp_data
                cort_wrkhrs_data_cell.value = 9.0
                cort_not_data_cell.value = float(wrkhrs_data) - 9.0
                cort_weekoff_data_cell.value = "P"
            elif row > 1:
                cort_compcode_data_cell.value = compcode_data
                cort_emp_data_cell.value = emp_data
                cort_date_data_cell.value = date_data
                cort_comp_data_cell.value = comp_data
                cort_wrkhrs_data_cell.value = wrkhrs_data
                cort_not_data_cell.value = not_data
                cort_weekoff_data_cell.value = "P"
            cort_row += 1


    # filtercell = "A1:T" + str(maxrowsh1 + 1)
    # sh3.auto_filter.ref = filtercell
    wb3.save(master.temppath + "exception.xlsx")
    wb4.save(master.temppath + "payroll.xlsx")
    status = "Done"
