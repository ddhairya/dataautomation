import datelist as dl
import master
import openpyxl as xl
from pathlib import Path
import numpy as np
from datetime import datetime

# we will get the missing dates from the file along with exceptional working hrs
step = "four"
status = ""

emp_date = []
emp = {}
if Path(master.temppath+"stepthree.xlsx").exists():
    print("Step " + step)
    wb = xl.load_workbook(master.tempfile+"stepthree.xlsx")
    sh1 = wb["stepthree"]
    sh2 = wb.create_sheet("stepfour")
    maxrow = sh1.max_row
    id = 1
    for row in range(1, maxrow+1):
        compcode_data = sh1.cell(row,1).value
        emp_data = sh1.cell(row,2).value
        date_data = sh1.cell(row,3).value
        comp_data = sh1.cell(row,4).value
        if emp_data is not None and date_data is not None:
            emp_date.append(date_data)
            # print(emp_date)
        elif emp_data is not None and date_data is None:
            if id == 1:
                emp_id = emp_data+compcode_data+comp_data
                # print(emp_id)
                id = 2
            elif id == 2:
                missing_date = np.setdiff1d(dl.datelist, emp_date).tolist()
                # print(missing_date)
                missing_date.sort(key=lambda date: datetime.strptime(date, "%d/%m/%Y"))

                emp[emp_id] = missing_date
                # print(emp[emp_id])
                id = 1
                emp_date = []
        elif emp_data is None and date_data is None:
            pass

    missing_compcode_emp_id = sh2.cell(1, 1)
    missing_date_emp_id = sh2.cell(1, 2)
    missing_date_emp_date = sh2.cell(1, 3)
    missing_comp = sh2.cell(1, 4)
    missing_wrkhr = sh2.cell(1, 5)
    missing_not = sh2.cell(1, 6)

    missing_compcode_emp_id.value = "cmpycode"
    missing_date_emp_id.value = "empcode"
    missing_date_emp_date.value = "SalaryDates"
    missing_comp.value = "Comp"
    missing_wrkhr.value = "Nhrs"
    missing_not.value = "N_ OT_HRS"
    noofrow = 2
    for key, values in emp.items():
        if (isinstance(values, list)):
            for value in values:
                # print(noofrow)
                missing_compcode_emp_id = sh2.cell(noofrow, 1)
                missing_date_emp_id = sh2.cell(noofrow, 2)
                missing_date_emp_date = sh2.cell(noofrow, 3)
                missing_comp = sh2.cell(noofrow, 4)

                missing_date_emp_id.value = key[:5]
                missing_compcode_emp_id.value = key[5:7]
                missing_date_emp_date.value = value
                missing_comp.value = key[7:]
                # print(key, value)
                noofrow += 1
        else:
            print("Error")
    wb.save(master.temppath+"stepfour.xlsx")
    status = "Done"