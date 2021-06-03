import master
import openpyxl as xl
from pathlib import Path

# combine the missing date and exceptional working hours.
step = "five"
status = ""

if Path(master.temppath+"stepfour.xlsx").exists():
    print("Step " + step)
    wb = xl.load_workbook(master.tempfile+"stepfour.xlsx")
    sh2 = wb["stepfour"]
    sh1 = wb["stepthree"]
    maxrowsh1 = sh1.max_row
    maxrowsh2 = sh2.max_row
    for row in range(1, maxrowsh1+1):
        compcode_data = sh1.cell(row,1).value
        emp_data = sh1.cell(row,2).value
        date_data = sh1.cell(row,3).value
        comp_data = sh1.cell(row,4).value
        wrkhrs_data = sh1.cell(row,5).value

        if wrkhrs_data is not None:
            miss_compcode_data_cell = sh2.cell(maxrowsh2, 1)
            miss_emp_data_cell = sh2.cell(maxrowsh2,2)
            miss_date_data_cell = sh2.cell(maxrowsh2,3)
            miss_comp_data_cell = sh2.cell(maxrowsh2,4)
            miss_wrkhrs_data_cell = sh2.cell(maxrowsh2,5)

            miss_compcode_data_cell.value = compcode_data
            miss_emp_data_cell.value = emp_data
            miss_date_data_cell.value = date_data
            miss_comp_data_cell.value = comp_data
            miss_wrkhrs_data_cell.value = wrkhrs_data
            maxrowsh2 += 1
    # filtercell = "A1:T" + str(maxrowsh2+1)
    # sh2.auto_filter.ref = filtercell
    wb.save(master.temppath+"stepfive.xlsx")
    status = "Done"
