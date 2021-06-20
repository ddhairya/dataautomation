import master
import openpyxl as xl
from pathlib import Path

# it will make two file one workbook with the correct employee data and a sheet with exceptions.
step = "Three"
status = ""

if Path(master.temppath+"steptwo.xlsx").exists():
    print("Step "+ step)
    wb1 = xl.load_workbook(master.tempfile+"steptwo.xlsx")
    sh1 = wb1["steptwo"]
    sh3 = wb1.create_sheet("stepthree")
    wb2 = xl.Workbook()
    sh2 = wb2["Sheet"]
    maxrow = sh1.max_row

    new_compcode_data_cell = sh2.cell(1, 1)
    new_emp_data_cell = sh2.cell(1, 2)
    new_date_data_cell = sh2.cell(1, 3)
    new_comp_data_cell = sh2.cell(1, 4)
    new_wrkhrs_data_cell = sh2.cell(1, 5)
    new_not_data_cell = sh2.cell(1, 6)

    new_compcode_data_cell.value = "cmpycode"
    new_emp_data_cell.value = "empcode"
    new_date_data_cell.value = "SalaryDates"
    new_comp_data_cell.value = "Comp"
    new_wrkhrs_data_cell.value = "Nhrs"
    new_not_data_cell.value = "N_ OT_HRS"
    for row in range(1, maxrow+1):
        compcode_data = sh1.cell(row,1).value
        emp_data = sh1.cell(row,2).value
        date_data = sh1.cell(row,3).value
        comp_data = sh1.cell(row,4).value
        wrkhrs_data = sh1.cell(row,5).value

        miss_compcode_data_cell = sh3.cell(row, 1)
        miss_emp_data_cell = sh3.cell(row, 2)
        miss_date_data_cell = sh3.cell(row, 3)
        miss_comp_data_cell = sh3.cell(row, 4)
        miss_wrkhrs_data_cell = sh3.cell(row, 5)

        new_compcode_data_cell = sh2.cell(row+1,1)
        new_emp_data_cell = sh2.cell(row+1,2)
        new_date_data_cell = sh2.cell(row+1, 3)
        new_comp_data_cell = sh2.cell(row+1, 4)
        new_wrkhrs_data_cell = sh2.cell(row+1, 5)

        if date_data is not None and isinstance(wrkhrs_data,float) and wrkhrs_data > master.max_work_hr_exception:
            miss_compcode_data_cell.value = compcode_data
            miss_emp_data_cell.value = emp_data[:master.emp]
            miss_date_data_cell.value = date_data
            miss_comp_data_cell.value = comp_data
            miss_wrkhrs_data_cell.value = wrkhrs_data

            # This is to create a record in the payroll file even working hrs are more than 11,
            # by avoiding actual work hrs and puting the static.

            # new_compcode_data_cell.value = compcode_data
            # new_emp_data_cell.value = emp_data[:master.emp]
            # new_date_data_cell.value = date_data
            # new_comp_data_cell.value = comp_data
            # new_wrkhrs_data_cell.value = 11

        elif date_data is not None and isinstance(wrkhrs_data,float) and wrkhrs_data < 0.0:
            miss_compcode_data_cell.value = compcode_data
            miss_emp_data_cell.value = emp_data[:master.emp]
            miss_date_data_cell.value = date_data
            miss_comp_data_cell.value = comp_data
            miss_wrkhrs_data_cell.value = wrkhrs_data
        elif comp_data is not None:
            miss_compcode_data_cell.value = compcode_data
            miss_emp_data_cell.value = emp_data[:master.emp]
            miss_date_data_cell.value = date_data
            miss_comp_data_cell.value = comp_data
            miss_wrkhrs_data_cell.value = None

            new_compcode_data_cell.value = compcode_data
            new_emp_data_cell.value = emp_data[:master.emp]
            new_date_data_cell.value = date_data
            new_comp_data_cell.value = comp_data
            new_wrkhrs_data_cell.value = wrkhrs_data


    wb1.save(master.temppath+"stepthree.xlsx")
    wb2.save(master.temppath+"payroll.xlsx")
    status = "Done"